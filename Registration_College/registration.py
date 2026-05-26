from tkinter import *
import sqlite3
from openpyxl import Workbook, load_workbook
import os

def registration(loc, frame1, go_home):   # ✅ added go_home

    frame1.destroy()

    frame2 = Frame(loc, bg='white')
    frame2.pack(expand=True)

    Heading = Label(
        frame2,
        text="Student's Data:",
        font=('Bookman Old Style', 30, 'bold'),
        bg='white',
        fg='#271d56'
    )
    Heading.pack(pady=20)

    data = Frame(frame2, bg='white')
    data.pack(pady=20)

    returns = {}

    entry_style = {
        "bg": "white",
        "fg": "black",
        "font": ('Bookman Old Style', 15),
        "bd": 2,
        "relief": "solid",
        "width": 25
    }

    # ---- VALIDATION (Marks only numbers) ----
    def only_numbers(P):
        if P == "":
            return True
        return P.isdigit() and int(P) <= 100

    vcmd = frame2.register(only_numbers)

    # ---- Name & Address ----
    fields = ['Name', 'Address']
    for i, field in enumerate(fields):

        Label(data, text=field, bg='white',
              font=('Bookman Old Style', 15)
        ).grid(row=i, column=0, sticky=W, pady=10, padx=15)

        entry = Entry(data, **entry_style)
        entry.grid(row=i, column=1, pady=10)

        returns[field] = entry

    # ---- DOB ----
    Label(data, text="DOB", bg='white',
          font=('Bookman Old Style', 15)
    ).grid(row=2, column=0, sticky=W, pady=10, padx=15)

    dob_frame = Frame(data, bg='white')
    dob_frame.grid(row=2, column=1, pady=10, sticky=W)

    Day = StringVar(value='Day')
    Month = StringVar(value='Month')
    Year = StringVar(value='Year')

    OptionMenu(dob_frame, Day, *range(1, 32)).pack(side=LEFT, padx=5)
    OptionMenu(dob_frame, Month, *range(1, 13)).pack(side=LEFT, padx=5)
    OptionMenu(dob_frame, Year, *range(1990, 2026)).pack(side=LEFT, padx=5)

    returns['Day'] = Day
    returns['Month'] = Month
    returns['Year'] = Year

    # ---- Department ----
    Label(data, text="Department", bg='white',
          font=('Bookman Old Style', 15)
    ).grid(row=3, column=0, sticky=W, pady=10, padx=15)

    dept_var = StringVar(value="Select Department")

    departments = [
        "Computer Science",
        "Mechanical",
        "Civil",
        "Electrical",
        "Electronics & Communication",
        "Information Technology"
    ]

    OptionMenu(data, dept_var, *departments).grid(row=3, column=1, pady=10)

    returns['Department'] = dept_var

    # ---- Marks ----
    marks_fields = ['10th %', '12th %']
    for i, field in enumerate(marks_fields, start=4):

        Label(data, text=field, bg='white',
              font=('Bookman Old Style', 15)
        ).grid(row=i, column=0, sticky=W, pady=10, padx=15)

        entry = Entry(
            data,
            **entry_style,
            validate="key",
            validatecommand=(vcmd, '%P')
        )
        entry.grid(row=i, column=1, pady=10)

        returns[field] = entry

    # ---- Gender ----
    Label(data, text="Gender", bg='white',
          font=('Bookman Old Style', 15)
    ).grid(row=6, column=0, sticky=W, pady=10, padx=15)

    gender_var = StringVar(value="")

    frame_gender = Frame(data, bg='white')
    frame_gender.grid(row=6, column=1)

    Radiobutton(frame_gender, text="Male", variable=gender_var, value="Male", bg='white').pack(side=LEFT, padx=5)
    Radiobutton(frame_gender, text="Female", variable=gender_var, value="Female", bg='white').pack(side=LEFT, padx=5)

    returns['Gender'] = gender_var

    # ---- Error Label ----
    error_label = Label(frame2, text="", fg="red", bg="white",
                        font=('Bookman Old Style', 12))
    error_label.pack()

    # ---- VALIDATE ----
    def validate():
        empty = []

        for k, v in returns.items():
            val = v.get()
            if val in ["", "Day", "Month", "Year", "Select Department"]:
                empty.append(k)

        if empty:
            error_label.config(text="Fill: " + ", ".join(empty))
            return False

        error_label.config(text="")
        return True

    # ---- DATABASE ----
    conn = sqlite3.connect("students.db")
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS students (
        name TEXT,
        address TEXT,
        dob TEXT,
        department TEXT,
        marks10 INTEGER,
        marks12 INTEGER,
        gender TEXT
    )
    """)
    conn.commit()
    conn.close()

    # ---- SUBMIT ----
    def submit():
        if not validate():
            return

        data_dict = {k: v.get() for k, v in returns.items()}

        # ---- Combine DOB ----
        dob = f"{data_dict['Day']}-{data_dict['Month']}-{data_dict['Year']}"

        # ---- SAVE TO DATABASE ----
        conn = sqlite3.connect("students.db")
        c = conn.cursor()

        c.execute("""
        INSERT INTO students VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            data_dict['Name'],
            data_dict['Address'],
            dob,
            data_dict['Department'],
            data_dict['10th %'],
            data_dict['12th %'],
            data_dict['Gender']
        ))

        conn.commit()
        conn.close()

        # ---- SAVE TO EXCEL ----
        file_name = "students.xlsx"

        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active

            # headers (only first time)
            ws.append([
                "Name", "Address", "DOB",
                "Department", "10th %", "12th %", "Gender"
            ])
        else:
            wb = load_workbook(file_name)
            ws = wb.active

        # add row
        ws.append([
            data_dict['Name'],
            data_dict['Address'],
            dob,
            data_dict['Department'],
            data_dict['10th %'],
            data_dict['12th %'],
            data_dict['Gender']
        ])

        wb.save(file_name)

        print("Saved to DB + Excel")

        # ---- GO BACK HOME ----
        frame2.destroy()
        go_home(loc)
        
    Button(
        frame2,
        text="Submit",
        bg="black",
        fg="white",
        font=('Bookman Old Style', 15),
        command=submit
    ).pack(pady=20)